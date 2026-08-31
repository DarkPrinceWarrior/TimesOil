"""Safe adapter for the organizers' official CHDD_PYTHON calculator."""

from __future__ import annotations

import asyncio
import csv
import hashlib
import json
import math
import os
import re
import subprocess
import sys
import tempfile
from collections.abc import Iterable, Mapping
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any
from zipfile import ZIP_DEFLATED, ZipFile, ZipInfo

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell

CHDD_FIELDS = (
    "DATA",
    "well",
    "WLPT",
    "WLPR",
    "WOMT",
    "WOMR",
    "WWIR",
    "WWIT",
    "THP",
    "BHP",
    "WEFF",
    "WLPT_Diff",
    "WOMT_Diff",
    "WWIT_Diff",
)
_NUMERIC_FIELDS = CHDD_FIELDS[2:]
_CALCULATOR_FILES = ("РАСЧЕТ_ЧДД.py", "chdd_model.py", "excel_io.py")
_VERSION_RE = re.compile(r'^VERSION\s*=\s*["\']([^"\']+)["\']', re.MULTILINE)


class EconomicsError(RuntimeError):
    """The official economics calculation could not be certified."""


@dataclass(frozen=True, slots=True)
class EconomicResult:
    total_chdd_m: float
    profitability_index: float
    start_date: str
    max_date: str
    diagnostics: Mapping[str, Any]
    output_dir: Path
    manifest_path: Path


class CHDDEconomicsAdapter:
    """Normalize one 14-field trajectory and invoke CHDD without a shell."""

    def __init__(
        self,
        chdd_python_dir: str | Path | None = None,
        *,
        timeout_seconds: float = 120.0,
        python_executable: str | Path = sys.executable,
    ) -> None:
        repository = Path(__file__).resolve().parents[3]
        self.chdd_dir = Path(
            chdd_python_dir
            or repository / "docs" / "hackathon" / "chdd" / "CHDD_PYTHON"
        ).resolve()
        self.timeout_seconds = timeout_seconds
        self.python_executable = Path(python_executable).resolve()
        if not 0 < timeout_seconds <= 3600:
            raise ValueError("CHDD timeout must be in (0, 3600]")
        if not self.python_executable.is_file():
            raise FileNotFoundError(f"Python executable not found: {self.python_executable}")
        missing = [name for name in _CALCULATOR_FILES if not (self.chdd_dir / name).is_file()]
        if missing:
            raise FileNotFoundError(f"CHDD_PYTHON is incomplete: {', '.join(missing)}")
        self.norms_path = self.chdd_dir / "input" / "Нормативы_ЧДД.xlsx"
        if not self.norms_path.is_file():
            raise FileNotFoundError(f"CHDD norms not found: {self.norms_path}")

    @classmethod
    def from_env(cls, environ: Mapping[str, str] | None = None) -> CHDDEconomicsAdapter:
        source = os.environ if environ is None else environ
        try:
            timeout = float(source.get("CHDD_TIMEOUT_SECONDS", "120"))
        except ValueError:
            raise ValueError("CHDD_TIMEOUT_SECONDS must be numeric") from None
        return cls(source.get("CHDD_PYTHON_DIR") or None, timeout_seconds=timeout)

    def calculate(
        self,
        records: Iterable[Mapping[str, Any]],
        *,
        start_year: int,
        output_dir: str | Path,
        charge_initial_pump: bool | None = None,
    ) -> EconomicResult:
        if isinstance(start_year, bool) or not isinstance(start_year, int) or not 1900 <= start_year <= 9999:
            raise ValueError("start_year must be an explicit four-digit year")
        if charge_initial_pump is not None and not isinstance(charge_initial_pump, bool):
            raise TypeError("charge_initial_pump must be bool or None")
        destination = Path(output_dir).resolve()
        if destination.exists():
            raise FileExistsError(f"CHDD run directory already exists: {destination}")
        rows = normalize_chdd_rows(records)
        if not any(int(str(row["DATA"])[:4]) >= start_year for row in rows):
            raise ValueError("CHDD input contains no records at or after start_year")
        destination.parent.mkdir(parents=True, exist_ok=True)

        with tempfile.TemporaryDirectory(prefix=".chdd-", dir=destination.parent) as temporary:
            work = Path(temporary)
            input_path = work / "input.csv"
            result_path = work / "result.json"
            report_path = work / "report.xlsx"
            effective_norms = self.norms_path
            overrides: dict[str, bool] = {}
            if charge_initial_pump is not None:
                effective_norms = work / "norms-effective.xlsx"
                _write_charge_initial_pump_norms(
                    self.norms_path, effective_norms, charge_initial_pump
                )
                overrides["chargeInitialPump"] = charge_initial_pump
            _write_csv(input_path, rows)
            command = [
                str(self.python_executable),
                str(self.chdd_dir / "РАСЧЕТ_ЧДД.py"),
                "--input",
                str(input_path),
                "--norms",
                str(effective_norms),
                "--start-year",
                str(start_year),
                "--output",
                str(report_path),
                "--json",
                str(result_path),
            ]
            try:
                completed = subprocess.run(
                    command,
                    cwd=self.chdd_dir,
                    env=_subprocess_env(),
                    stdin=subprocess.DEVNULL,
                    capture_output=True,
                    text=True,
                    encoding="utf-8",
                    errors="replace",
                    timeout=self.timeout_seconds,
                    check=False,
                    shell=False,
                )
            except subprocess.TimeoutExpired as exc:
                raise EconomicsError("official CHDD calculation timed out") from exc
            if completed.returncode != 0:
                raise EconomicsError(
                    f"official CHDD calculation failed with exit code {completed.returncode}"
                )
            if not result_path.is_file() or not report_path.is_file():
                raise EconomicsError("official CHDD calculation did not produce required artifacts")
            try:
                raw_result = json.loads(result_path.read_text(encoding="utf-8"))
            except (OSError, json.JSONDecodeError) as exc:
                raise EconomicsError("official CHDD result JSON is invalid") from exc
            result = _validated_result(raw_result, expected_start_year=start_year)
            manifest = _manifest(
                adapter=self,
                rows=rows,
                start_year=start_year,
                input_path=input_path,
                result_path=result_path,
                raw_result=raw_result,
                effective_norms=effective_norms,
                assumption_overrides=overrides,
            )
            (work / "manifest.json").write_text(
                json.dumps(manifest, ensure_ascii=False, indent=2, sort_keys=True),
                encoding="utf-8",
            )
            work.rename(destination)

        return EconomicResult(
            total_chdd_m=result["total_chdd_m"],
            profitability_index=result["profitability_index"],
            start_date=result["start_date"],
            max_date=result["max_date"],
            diagnostics=result["diagnostics"],
            output_dir=destination,
            manifest_path=destination / "manifest.json",
        )

    async def calculate_async(
        self,
        records: Iterable[Mapping[str, Any]],
        *,
        start_year: int,
        output_dir: str | Path,
        charge_initial_pump: bool | None = None,
    ) -> EconomicResult:
        materialized = list(records)
        return await asyncio.to_thread(
            self.calculate,
            materialized,
            start_year=start_year,
            output_dir=output_dir,
            charge_initial_pump=charge_initial_pump,
        )


def normalize_chdd_rows(records: Iterable[Mapping[str, Any]]) -> list[dict[str, str | float]]:
    rows: list[dict[str, str | float]] = []
    seen: set[tuple[str, str]] = set()
    for index, record in enumerate(records):
        missing = set(map(str, CHDD_FIELDS)) - set(record)
        if missing:
            raise ValueError(f"CHDD row {index} misses fields: {sorted(missing)}")
        normalized_date = _date(record["DATA"], row=index)
        well = str(record["well"]).strip()
        if (
            not well
            or len(well) > 128
            or any(ord(character) < 32 for character in well)
            or well[0] in "=+-@"
        ):
            raise ValueError(f"CHDD row {index} has an invalid well")
        duplicate_key = (normalized_date[:7], well)
        if duplicate_key in seen:
            raise ValueError(f"duplicate CHDD well-month: {well} {normalized_date[:7]}")
        seen.add(duplicate_key)
        row: dict[str, str | float] = {"DATA": normalized_date, "well": well}
        for field in _NUMERIC_FIELDS:
            row[field] = _number(record[field], field=field, row=index)
        rows.append(row)
    if not rows:
        raise ValueError("CHDD input must contain at least one row")
    return sorted(rows, key=lambda row: (str(row["DATA"]), str(row["well"])))


def _date(value: Any, *, row: int) -> str:
    if isinstance(value, datetime):
        parsed = value.date()
    elif isinstance(value, date):
        parsed = value
    else:
        text = str(value).strip()
        try:
            parts = [int(item) for item in re.split(r"[-./]", text)]
            if len(parts) != 3:
                raise ValueError
            year, month, day = parts if len(str(parts[0])) == 4 else (parts[2], parts[1], parts[0])
            parsed = date(year, month, day)
        except (TypeError, ValueError):
            raise ValueError(f"CHDD row {row} has an invalid DATA")
    return parsed.isoformat()


def _number(value: Any, *, field: str, row: int) -> float:
    if isinstance(value, bool):
        raise TypeError(f"CHDD row {row} field {field} must be numeric")
    try:
        number = float(str(value).strip().replace("\u00a0", "").replace(" ", "").replace(",", "."))
    except (TypeError, ValueError):
        raise ValueError(f"CHDD row {row} field {field} must be numeric") from None
    if not math.isfinite(number):
        raise ValueError(f"CHDD row {row} field {field} must be finite")
    return number


def _write_csv(path: Path, rows: list[dict[str, str | float]]) -> None:
    with path.open("w", encoding="utf-8", newline="") as stream:
        fieldnames: list[str] = list(map(str, CHDD_FIELDS))
        writer: csv.DictWriter[str] = csv.DictWriter(
            stream, fieldnames=fieldnames, extrasaction="ignore"
        )
        writer.writeheader()
        writer.writerows(rows)


def _write_charge_initial_pump_norms(source: Path, destination: Path, value: bool) -> None:
    workbook = load_workbook(source)
    matches = [
        row[2]
        for sheet in workbook.worksheets
        for row in sheet.iter_rows()
        if len(row) >= 3
        and isinstance(row[2], Cell)
        and str(row[0].value or "").strip() == "chargeInitialPump"
    ]
    if len(matches) != 1:
        raise EconomicsError("CHDD norms lack a unique chargeInitialPump row")
    matches[0].value = "ДА" if value else "НЕТ"
    workbook.properties.created = datetime(2000, 1, 1)
    workbook.properties.modified = datetime(2000, 1, 1)
    workbook.save(destination)
    workbook.close()
    _canonicalize_xlsx(destination)


def _canonicalize_xlsx(path: Path) -> None:
    with ZipFile(path) as source:
        entries = []
        for name in sorted(source.namelist()):
            data = source.read(name)
            if name == "docProps/core.xml":
                data = re.sub(
                    rb"(<dcterms:modified\b[^>]*>)[^<]*(</dcterms:modified>)",
                    rb"\g<1>2000-01-01T00:00:00Z\g<2>",
                    data,
                )
            entries.append((name, data, source.getinfo(name).external_attr))
    temporary = path.with_name(f".{path.name}.canonical")
    with ZipFile(temporary, "w", compression=ZIP_DEFLATED, compresslevel=9) as target:
        for name, data, external_attr in entries:
            info = ZipInfo(name, date_time=(1980, 1, 1, 0, 0, 0))
            info.compress_type = ZIP_DEFLATED
            info.external_attr = external_attr
            target.writestr(info, data)
    temporary.replace(path)


def _validated_result(raw: Any, *, expected_start_year: int) -> dict[str, Any]:
    if not isinstance(raw, Mapping) or not isinstance(raw.get("summary"), Mapping):
        raise EconomicsError("official CHDD result lacks a summary")
    start_date = raw.get("startDate")
    max_date = raw.get("maxDate")
    diagnostics = raw.get("diagnostics", {})
    if (
        not isinstance(start_date, str)
        or not start_date.startswith(f"{expected_start_year:04d}-")
        or not isinstance(max_date, str)
        or not isinstance(diagnostics, Mapping)
    ):
        raise EconomicsError("official CHDD result has invalid provenance fields")
    summary = raw["summary"]
    total = _finite_result(summary.get("totalChddM"), "totalChddM")
    index = _finite_result(summary.get("profitabilityIndex"), "profitabilityIndex")
    return {
        "total_chdd_m": total,
        "profitability_index": index,
        "start_date": start_date,
        "max_date": max_date,
        "diagnostics": dict(diagnostics),
    }


def _finite_result(value: Any, name: str) -> float:
    if isinstance(value, bool):
        raise EconomicsError(f"official CHDD {name} is invalid")
    try:
        number = float(value)
    except (TypeError, ValueError):
        raise EconomicsError(f"official CHDD {name} is invalid") from None
    if not math.isfinite(number):
        raise EconomicsError(f"official CHDD {name} is not finite")
    return number


def _manifest(
    *,
    adapter: CHDDEconomicsAdapter,
    rows: list[dict[str, str | float]],
    start_year: int,
    input_path: Path,
    result_path: Path,
    raw_result: Mapping[str, Any],
    effective_norms: Path,
    assumption_overrides: Mapping[str, bool],
) -> dict[str, Any]:
    calculator_hashes = {
        name: _sha256(adapter.chdd_dir / name) for name in _CALCULATOR_FILES
    }
    model_source = (adapter.chdd_dir / "chdd_model.py").read_text(encoding="utf-8")
    version_match = _VERSION_RE.search(model_source)
    return {
        "schema_version": 1,
        "adapter": "timesoil.aios.CHDDEconomicsAdapter",
        "start_year": start_year,
        "fields": list(CHDD_FIELDS),
        "row_count": len(rows),
        "input_sha256": _sha256(input_path),
        "norms_source_sha256": _sha256(adapter.norms_path),
        "norms_sha256": _sha256(effective_norms),
        "assumption_overrides": dict(assumption_overrides),
        "calculator_sha256": calculator_hashes,
        "calculator_version": version_match.group(1) if version_match else "unknown",
        "result_sha256": _sha256(result_path),
        "artifacts": {
            "input": "input.csv",
            "result": "result.json",
            "report": "report.xlsx",
            **(
                {"effective_norms": effective_norms.name}
                if assumption_overrides
                else {}
            ),
        },
        "summary": raw_result.get("summary", {}),
    }


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _subprocess_env() -> dict[str, str]:
    environment = {
        "PATH": os.environ.get("PATH", os.defpath),
        "LANG": os.environ.get("LANG", "C.UTF-8"),
        "LC_ALL": os.environ.get("LC_ALL", "C.UTF-8"),
        "PYTHONHASHSEED": "0",
        "PYTHONIOENCODING": "utf-8",
    }
    import_paths = [str(Path(item).resolve()) for item in sys.path if item and Path(item).is_dir()]
    if import_paths:
        environment["PYTHONPATH"] = os.pathsep.join(dict.fromkeys(import_paths))
    return environment
