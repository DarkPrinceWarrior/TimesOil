#!/usr/bin/env python3
"""Run the isolated A100 Model Y baseline and six-month Track 1 proof."""

from __future__ import annotations

import argparse
import csv
from dataclasses import replace
from datetime import date
from hashlib import sha256
import json
from math import isclose
import os
from pathlib import Path
import stat
import subprocess
import sys
from uuid import uuid4

from openpyxl import load_workbook

from timesoil.aios.contracts import (
    Case,
    ControlAction,
    ControlTarget,
    State,
    WellRole,
    WellState,
    WellStatus,
)
from timesoil.aios.economics import CHDDEconomicsAdapter
from timesoil.aios.opm import (
    OPM_EXPORT_VECTORS,
    OpmFlowRunner,
    OpmGdmBackend,
    OpmRunResult,
    _sha256_file,
    verify_summary_extraction,
)
from timesoil.aios.opm_chdd import export_opm_chdd
from timesoil.aios.schedule import ScheduleCompiler


ROOT = Path("/tmp/timesoil-kt2/track1-v1")
SOURCE = Path("/tmp/timesoil-kt2/model_y/Model_Y (3).zip")
SOURCE_SHA256 = "261591b458084eaaf8c86a601e68d3bdc6e91fed9f0117fdcbe58cfca4eb882e"
BASELINE = ROOT / "results/model-y-baseline-20260831-a100-v4"
CANDIDATES = ROOT / "results/model-y-track1-candidates-20260831-a100-v3"
FINAL = ROOT / "results/model-y-track1-mpc-20260831-a100-v3"
CONFIG = ROOT / "config/model-y-track1-proof.json"
ORGANIZER_REFERENCE = (
    ROOT
    / "docs/hackathon/chdd/reference_baselines"
    / "Расчет ЧДД через OPM Flow Model_Y.xlsx"
)
DECK = "MODEL_Y/MODEL_Y.DATA"
SCHEDULE = "MODEL_Y/INCLUDE/DemoSpe_002_2_sch.inc"
TARGETS = {2007: 5326.453465501771, 2014: 1082.233695354093}
TARGET_PROFILES = {2007: "organizer_reference", 2014: "canonical"}
SIMULATOR_ENVELOPE_2007 = (5261.158803096247, 5326.453465501771)
CHDD_ABS_TOLERANCE_M = SIMULATOR_ENVELOPE_2007[1] - SIMULATOR_ENVELOPE_2007[0]
MASS_VECTORS = frozenset(OPM_EXPORT_VECTORS)
_SUMMARY_VECTOR_FIELDS = {"quantity", "unit", "chdd_field", "transform"}
HORIZON_START = date(2014, 1, 1)
HORIZON_END = date(2014, 6, 1)
HORIZON_MONTH_COUNT = 6
CONTROL_PERTURBATION = 10.0


def _next_month(month: date) -> date:
    return date(month.year + (month.month == 12), month.month % 12 + 1, 1)


def _month_range(start: date, end: date) -> tuple[date, ...]:
    months: list[date] = []
    month = start
    while month <= end:
        months.append(month)
        month = _next_month(month)
    return tuple(months)


def _validated_manifest_vectors(summary_contract: object) -> frozenset[str]:
    if not isinstance(summary_contract, dict):
        raise RuntimeError("OPM run manifest summary vector contract is invalid")
    vector_values = summary_contract.get("vectors")
    if not isinstance(vector_values, dict) or not vector_values or any(
        not isinstance(name, str)
        or not name
        or not isinstance(metadata, dict)
        or set(metadata) != _SUMMARY_VECTOR_FIELDS
        or not all(isinstance(value, str) for value in metadata.values())
        for name, metadata in vector_values.items()
    ):
        raise RuntimeError("OPM run manifest summary vector contract is invalid")
    return frozenset(vector_values)


def _write_new_json(path: Path, value: object) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("x", encoding="utf-8") as stream:
        json.dump(value, stream, ensure_ascii=False, sort_keys=True, indent=2)
        stream.write("\n")


def _assert_no_running_opm(*, _run=subprocess.run) -> None:
    completed = _run(
        ["docker", "ps", "--format", "{{.Image}} {{.Names}}"],
        stdin=subprocess.DEVNULL,
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
        check=False,
        shell=False,
    )
    if completed.returncode != 0:
        raise RuntimeError("cannot verify running Docker containers")
    conflicts = [
        line
        for line in completed.stdout.splitlines()
        if "opmreleases:2026.04" in line
    ]
    if conflicts:
        raise RuntimeError(f"another OPM 2026.04 container is running: {conflicts}")


def _proof_source_contract(controller: Path) -> dict[str, object]:
    proof_script = Path(__file__).absolute()
    controller_source = proof_script.with_name("run_track1_mpc.py")
    paths = (proof_script, controller_source, controller.absolute())
    for path in paths:
        if path.resolve() != path or path.is_symlink() or not path.is_file():
            raise RuntimeError(f"script provenance path is not a regular file: {path}")
    source_sha256 = _sha256_file(controller_source)
    executed_sha256 = _sha256_file(controller)
    if source_sha256 != executed_sha256:
        raise RuntimeError("executed run_track1_mpc.py differs from source bytes")
    executes_sibling = controller.absolute() == controller_source
    return {
        "execution": {
            "run_model_y_track1_proof.py": {
                "path": str(proof_script),
                "sha256": _sha256_file(proof_script),
            },
            "run_track1_mpc.py": {
                "path": str(controller.absolute()),
                "sha256": executed_sha256,
            },
        },
        "controller_source": {
            "path": str(controller_source),
            "sha256": source_sha256,
        },
        "controller_execution_mode": (
            "current-source-sibling" if executes_sibling else "staged-byte-copy"
        ),
        "copied_controller_bytes_exact": not executes_sibling,
    }


def _verify_proof_source_contract(contract: dict[str, object]) -> None:
    try:
        execution = contract["execution"]
        if not isinstance(execution, dict):
            raise TypeError
        controller = execution["run_track1_mpc.py"]
        if not isinstance(controller, dict) or not isinstance(controller.get("path"), str):
            raise TypeError
        current = _proof_source_contract(Path(controller["path"]))
    except (KeyError, TypeError) as exc:
        raise RuntimeError("Track 1 proof script source contract is invalid") from exc
    if current != contract:
        raise RuntimeError("Track 1 proof script source changed during execution")


def _reject_symlink_components(path: Path) -> None:
    absolute = path.absolute()
    current = Path(absolute.anchor)
    for part in absolute.parts[1:]:
        current /= part
        if current.is_symlink():
            raise RuntimeError(f"symlink path component is forbidden: {current}")


def _paths_overlap(left: Path, right: Path) -> bool:
    return left == right or left in right.parents or right in left.parents


def _output_paths(
    output_root: Path | None,
    *,
    resume_baseline: bool,
    source: Path | None = None,
    baseline_dir: Path | None = None,
    reference_workbook: Path | None = None,
) -> tuple[Path, Path, Path, Path, Path]:
    source = SOURCE if source is None else source
    baseline_dir = BASELINE if baseline_dir is None else baseline_dir
    reference_workbook = (
        ORGANIZER_REFERENCE if reference_workbook is None else reference_workbook
    )
    generated = output_root is None
    requested = (
        output_root
        if output_root is not None
        else ROOT.parents[1] / f"timesoil-track1-{uuid4().hex}"
    )
    _reject_symlink_components(requested)
    root = requested.absolute().resolve()
    protected = {
        "staged controller tree": (ROOT / "scripts").absolute().resolve(),
        "immutable baseline": baseline_dir.absolute().resolve(),
        "Model Y source": source.absolute().resolve(),
        "organizer reference workbook": reference_workbook.absolute().resolve(),
        "proof source tree": Path(__file__).absolute().parents[1].resolve(),
    }
    for label, path in protected.items():
        if _paths_overlap(root, path):
            raise RuntimeError(f"output root overlaps {label}: {path}")
    if generated and root.exists():
        raise RuntimeError(f"generated output root already exists: {root}")
    if root.exists() and not root.is_dir():
        raise RuntimeError(f"output root is not a directory: {root}")
    baseline = (
        baseline_dir
        if resume_baseline
        else root / "results/model-y-baseline-20260831-a100-v4"
    )
    artifacts = (
        root / "results/model-y-baseline-resume-20260831-a100-v4"
        if resume_baseline
        else baseline
    )
    return (
        baseline,
        root / "results/model-y-track1-candidates-20260831-a100-v3",
        root / "results/model-y-track1-mpc-20260831-a100-v3",
        root / "config/model-y-track1-proof.json",
        artifacts,
    )


FileSnapshot = tuple[bytes, tuple[int, ...]]


def _stat_signature(path: Path) -> tuple[int, ...]:
    value = path.stat(follow_symlinks=False)
    return (
        value.st_dev,
        value.st_ino,
        value.st_mode,
        value.st_size,
        value.st_mtime_ns,
        value.st_ctime_ns,
    )


def _snapshot_regular_file(path: Path, label: str) -> FileSnapshot:
    _reject_symlink_components(path)
    try:
        descriptor = os.open(path, os.O_RDONLY | os.O_CLOEXEC | os.O_NOFOLLOW)
    except OSError as exc:
        raise RuntimeError(f"{label} is not a regular non-symlink file: {path}") from exc
    try:
        before_stat = os.fstat(descriptor)
        if not stat.S_ISREG(before_stat.st_mode):
            raise RuntimeError(f"{label} is not a regular non-symlink file: {path}")
        before = (
            before_stat.st_dev,
            before_stat.st_ino,
            before_stat.st_mode,
            before_stat.st_size,
            before_stat.st_mtime_ns,
            before_stat.st_ctime_ns,
        )
        with os.fdopen(descriptor, "rb", closefd=False) as stream:
            value = stream.read()
        after_stat = os.fstat(descriptor)
        after = (
            after_stat.st_dev,
            after_stat.st_ino,
            after_stat.st_mode,
            after_stat.st_size,
            after_stat.st_mtime_ns,
            after_stat.st_ctime_ns,
        )
    finally:
        os.close(descriptor)
    _reject_symlink_components(path)
    if path.is_symlink() or not path.is_file() or _stat_signature(path) != after:
        raise RuntimeError(f"{label} is not a regular non-symlink file: {path}")
    if before != after or len(value) != before[3]:
        raise RuntimeError(f"{label} changed while it was read")
    return value, after


def _read_regular_bytes(path: Path, label: str) -> bytes:
    return _snapshot_regular_file(path, label)[0]


def _runtime_regular_file(path: Path, label: str) -> Path:
    absolute = path.absolute()
    _reject_symlink_components(absolute)
    if absolute.resolve() != absolute or absolute.is_symlink() or not absolute.is_file():
        raise RuntimeError(f"{label} is not a regular non-symlink file: {absolute}")
    return absolute


def _safe_relative_path(value: str, label: str) -> str:
    path = Path(value)
    if path.is_absolute() or ".." in path.parts or "\\" in value or not value:
        raise RuntimeError(f"{label} must be a safe relative path")
    return path.as_posix()


def _unique_json_object(pairs: list[tuple[str, object]]) -> dict[str, object]:
    value: dict[str, object] = {}
    for key, item in pairs:
        if key in value:
            raise ValueError(f"duplicate JSON key: {key}")
        value[key] = item
    return value


def _json_bytes(value: bytes, label: str) -> dict[str, object]:
    def reject_constant(constant: str) -> object:
        raise ValueError(f"non-finite JSON constant: {constant}")

    try:
        result = json.loads(
            value,
            object_pairs_hook=_unique_json_object,
            parse_constant=reject_constant,
        )
    except (UnicodeError, json.JSONDecodeError, ValueError) as exc:
        raise RuntimeError(f"{label} is not unique-key UTF-8 JSON") from exc
    if not isinstance(result, dict):
        raise RuntimeError(f"{label} must be a JSON object")
    return result


def _authenticated_json(
    path: Path, expected_sha256: str, label: str
) -> tuple[dict[str, object], bytes]:
    value = _read_regular_bytes(path, label)
    if sha256(value).hexdigest() != expected_sha256:
        raise RuntimeError(f"{label} SHA-256 mismatch")
    return _json_bytes(value, label), value


def _assert_file_bytes(path: Path, expected: bytes, label: str) -> None:
    if _read_regular_bytes(path, label) != expected:
        raise RuntimeError(f"{label} changed during verification")


def _assert_file_snapshot(path: Path, expected: FileSnapshot, label: str) -> None:
    if _snapshot_regular_file(path, label) != expected:
        raise RuntimeError(f"{label} bytes or stat changed during verification")


def _resume_baseline(
    runner: OpmFlowRunner,
    *,
    baseline: Path | None = None,
    candidates: Path | None = None,
    source: Path | None = None,
    deck: str | None = None,
    schedule_relative_path: str | None = None,
) -> tuple[OpmRunResult, Path, Path]:
    baseline = BASELINE if baseline is None else baseline
    candidates = CANDIDATES if candidates is None else candidates
    source = SOURCE if source is None else source
    deck = DECK if deck is None else deck
    schedule_relative_path = (
        SCHEDULE if schedule_relative_path is None else schedule_relative_path
    )
    manifest_path = baseline / "manifest.json"
    sidecar = baseline / "manifest.sha256"
    manifest_bytes = _read_regular_bytes(manifest_path, "baseline manifest")
    manifest_sha = sha256(manifest_bytes).hexdigest()
    sidecar_bytes = _read_regular_bytes(sidecar, "baseline manifest sidecar")
    if sidecar_bytes != f"{manifest_sha}  manifest.json\n".encode("ascii"):
        raise RuntimeError("baseline manifest SHA-256 sidecar is invalid")
    manifest = _json_bytes(manifest_bytes, "baseline manifest")
    verifier = OpmGdmBackend(
        runner,
        source,
        runs_dir=candidates,
        deck=deck,
        schedule_include=schedule_relative_path,
        normalize_model_y=True,
        parsing_strictness="low",
        source_model="Model Y",
    )
    verifier._verify_opm_manifest(manifest_path, baseline=True)
    _assert_file_bytes(manifest_path, manifest_bytes, "baseline manifest")
    _assert_file_bytes(sidecar, sidecar_bytes, "baseline manifest sidecar")
    manifest_deck = manifest.get("deck")
    summary_contract = manifest.get("summary_contract")
    command = manifest.get("command")
    warnings = manifest.get("warnings")
    if (
        not isinstance(manifest_deck, str)
        or not isinstance(summary_contract, dict)
        or not isinstance(summary_contract.get("overlay"), str)
        or not isinstance(command, list)
        or not all(isinstance(item, str) for item in command)
        or not isinstance(warnings, list)
        or not all(isinstance(item, str) for item in warnings)
    ):
        raise RuntimeError("baseline manifest execution fields are invalid")
    result = OpmRunResult(
        baseline,
        baseline / "output",
        baseline / "input" / manifest_deck,
        baseline / "input" / summary_contract["overlay"],
        baseline / "stdout.log",
        baseline / "stderr.log",
        manifest_path,
        manifest_sha,
        tuple(command),
        tuple(warnings),
    )
    report = baseline / "summary-report.txt"
    extraction = baseline / "summary-extraction.json"
    verify_summary_extraction(
        report,
        extraction,
        manifest_path,
        _docker_executable=runner.docker_executable,
    )
    _assert_file_bytes(manifest_path, manifest_bytes, "baseline manifest")
    _assert_file_bytes(sidecar, sidecar_bytes, "baseline manifest sidecar")
    return result, report, extraction


def _lower_sha256(value: object, label: str) -> str:
    if (
        not isinstance(value, str)
        or len(value) != 64
        or any(character not in "0123456789abcdef" for character in value)
    ):
        raise RuntimeError(f"{label} must be a lowercase SHA-256 digest")
    return value


def _verified_manifest_artifact(
    run_dir: Path, value: object, expected_path: str, label: str
) -> FileSnapshot:
    if not isinstance(value, dict) or set(value) != {"path", "bytes", "sha256"}:
        raise RuntimeError(f"Track 1 {label} artifact contract is invalid")
    size = value["bytes"]
    if value["path"] != expected_path or isinstance(size, bool) or not isinstance(size, int):
        raise RuntimeError(f"Track 1 {label} artifact metadata is invalid")
    digest = _lower_sha256(value["sha256"], f"Track 1 {label}")
    snapshot = _snapshot_regular_file(run_dir / expected_path, f"Track 1 {label}")
    data = snapshot[0]
    if len(data) != size or sha256(data).hexdigest() != digest:
        raise RuntimeError(f"Track 1 {label} artifact hash mismatch")
    return snapshot


def _verify_track1_run(
    run_dir: Path, expected_source_contract: object
) -> tuple[dict[str, object], str, str, dict[Path, FileSnapshot]]:
    manifest_path = run_dir / "manifest.json"
    manifest_snapshot = _snapshot_regular_file(manifest_path, "Track 1 manifest")
    manifest_bytes = manifest_snapshot[0]
    manifest_sha256 = sha256(manifest_bytes).hexdigest()
    sidecar_path = run_dir / "manifest.sha256"
    sidecar_snapshot = _snapshot_regular_file(
        sidecar_path, "Track 1 manifest sidecar"
    )
    sidecar_bytes = sidecar_snapshot[0]
    if sidecar_bytes != f"{manifest_sha256}  manifest.json\n".encode("ascii"):
        raise RuntimeError("Track 1 manifest SHA-256 sidecar is invalid")
    manifest = _json_bytes(manifest_bytes, "Track 1 manifest")
    required_manifest = {
        "schema",
        "run_id",
        "input_sha256",
        "config_sha256",
        "source_sha256",
        "script_source_contract",
        "backend_provenance",
        "schedule_sha256",
        "trajectory_run_ids",
        "artifacts",
    }
    if (
        set(manifest) != required_manifest
        or manifest.get("schema") != "timesoil.aios.track1-mpc-manifest/v1"
    ):
        raise RuntimeError("Track 1 manifest schema or fields are invalid")
    for field in ("input_sha256", "config_sha256", "source_sha256", "schedule_sha256"):
        _lower_sha256(manifest[field], f"Track 1 manifest {field}")
    if not isinstance(manifest.get("run_id"), str) or manifest["run_id"] != run_dir.name:
        raise RuntimeError("Track 1 manifest run_id differs from result directory")
    if manifest.get("script_source_contract") != expected_source_contract:
        raise RuntimeError("Track 1 manifest script provenance differs from executed bytes")
    artifacts = manifest.get("artifacts")
    if not isinstance(artifacts, dict) or set(artifacts) != {"result", "schedule"}:
        raise RuntimeError("Track 1 manifest artifact set is invalid")
    result_snapshot = _verified_manifest_artifact(
        run_dir, artifacts["result"], "result.json", "result"
    )
    schedule_snapshot = _verified_manifest_artifact(
        run_dir, artifacts["schedule"], "wells_schedule.inc", "schedule"
    )
    result_bytes = result_snapshot[0]
    schedule_bytes = schedule_snapshot[0]
    result = _json_bytes(result_bytes, "Track 1 result")
    required_result = {
        "schema",
        "run_id",
        "input_sha256",
        "config_sha256",
        "source_sha256",
        "script_source_contract",
        "case_id",
        "schedule",
        "evidence",
    }
    if (
        set(result) != required_result
        or result.get("schema") != "timesoil.aios.track1-mpc-result/v1"
    ):
        raise RuntimeError("Track 1 result schema or fields are invalid")
    for field in (
        "run_id",
        "input_sha256",
        "config_sha256",
        "source_sha256",
        "script_source_contract",
    ):
        if result.get(field) != manifest.get(field):
            raise RuntimeError(f"Track 1 result {field} differs from manifest")
    schedule = result.get("schedule")
    evidence = result.get("evidence")
    if (
        not isinstance(schedule, dict)
        or set(schedule) != {"sha256", "text", "actions"}
        or not isinstance(schedule.get("actions"), list)
        or not isinstance(evidence, dict)
        or set(evidence) != {"backend_provenance", "trajectories", "step_economics"}
    ):
        raise RuntimeError("Track 1 result schedule or evidence is invalid")
    schedule_text = schedule.get("text")
    schedule_sha256 = _lower_sha256(schedule.get("sha256"), "Track 1 result schedule")
    if (
        not isinstance(schedule_text, str)
        or schedule_text.encode("utf-8") != schedule_bytes
        or schedule_sha256 != manifest["schedule_sha256"]
        or sha256(schedule_bytes).hexdigest() != schedule_sha256
    ):
        raise RuntimeError("Track 1 result schedule differs from manifest artifact")
    if evidence.get("backend_provenance") != manifest.get("backend_provenance"):
        raise RuntimeError("Track 1 backend provenance differs from manifest")
    trajectories = evidence.get("trajectories")
    economics = evidence.get("step_economics")
    trajectory_ids = manifest.get("trajectory_run_ids")
    if (
        not isinstance(trajectories, list)
        or not trajectories
        or not all(isinstance(item, dict) for item in trajectories)
        or not isinstance(economics, list)
        or len(economics) != len(trajectories)
        or not all(isinstance(item, dict) for item in economics)
        or trajectory_ids != [item.get("run_id") for item in trajectories]
        or trajectory_ids != [item.get("run_id") for item in economics]
    ):
        raise RuntimeError("Track 1 trajectory/economics lineage is inconsistent")
    trajectory_fields = {
        "run_id",
        "month",
        "simulator",
        "certified",
        "chdd_complete",
        "invariant_violations",
        "actions",
        "next_state",
    }
    economics_fields = {"run_id", "start_date", "npv_million_rub", "complete"}
    if any(
        set(item) != trajectory_fields
        or item.get("certified") is not True
        or item.get("chdd_complete") is not True
        or not isinstance(item.get("next_state"), dict)
        for item in trajectories
    ) or any(
        set(item) != economics_fields
        or item.get("complete") is not True
        or isinstance(item.get("npv_million_rub"), bool)
        or not isinstance(item.get("npv_million_rub"), (int, float))
        for item in economics
    ):
        raise RuntimeError("Track 1 trajectory/economics certification is invalid")
    snapshots = {
        manifest_path: manifest_snapshot,
        sidecar_path: sidecar_snapshot,
        run_dir / "result.json": result_snapshot,
        run_dir / "wells_schedule.inc": schedule_snapshot,
    }
    for path, snapshot in snapshots.items():
        _assert_file_snapshot(path, snapshot, f"Track 1 {path.name}")
    return result, manifest_sha256, schedule_sha256, snapshots


def _verified_selected_lineage(
    candidates: Path, selected_run_id: object, restart_ref: object
) -> tuple[Path, str, bytes, dict[Path, FileSnapshot]]:
    if (
        not isinstance(selected_run_id, str)
        or not selected_run_id
        or selected_run_id in {".", ".."}
        or "\\" in selected_run_id
        or len(Path(selected_run_id).parts) != 1
    ):
        raise RuntimeError("selected Track 1 run_id is unsafe")
    _reject_symlink_components(candidates)
    candidates = candidates.absolute().resolve()
    selected = candidates / selected_run_id
    _reject_symlink_components(selected)
    if selected.is_symlink() or not selected.is_dir():
        raise RuntimeError("selected Track 1 candidate directory is invalid")
    if not isinstance(restart_ref, str) or restart_ref.count("#sha256=") != 1:
        raise RuntimeError("selected Track 1 restart_ref is not SHA-256 authenticated")
    path_text, digest_value = restart_ref.split("#sha256=", 1)
    lineage_sha256 = _lower_sha256(digest_value, "selected Track 1 lineage")
    lineage_path = Path(path_text)
    _reject_symlink_components(lineage_path)
    if not lineage_path.is_absolute() or lineage_path.resolve() != selected / "lineage.json":
        raise RuntimeError("selected Track 1 lineage path differs from candidate")
    lineage_snapshot = _snapshot_regular_file(
        lineage_path, "selected Track 1 lineage"
    )
    lineage_bytes = lineage_snapshot[0]
    if sha256(lineage_bytes).hexdigest() != lineage_sha256:
        raise RuntimeError("selected Track 1 lineage SHA-256 mismatch")
    sidecar = lineage_path.with_suffix(".sha256")
    sidecar_snapshot = _snapshot_regular_file(
        sidecar, "selected Track 1 lineage sidecar"
    )
    sidecar_bytes = sidecar_snapshot[0]
    if sidecar_bytes != f"{lineage_sha256}  {lineage_path.name}\n".encode("ascii"):
        raise RuntimeError("selected Track 1 lineage sidecar is invalid")
    lineage = _json_bytes(lineage_bytes, "selected Track 1 lineage")
    if (
        lineage.get("schema") != "timesoil.aios.track1-opm-lineage/v1"
        or lineage.get("status") != "certified"
        or lineage.get("run_id") != selected_run_id
    ):
        raise RuntimeError("selected Track 1 lineage contract is invalid")
    artifacts = lineage.get("artifacts")
    if (
        not isinstance(artifacts, list)
        or any(
            not isinstance(item, dict) or set(item) != {"purpose", "path", "sha256"}
            for item in artifacts
        )
    ):
        raise RuntimeError("selected Track 1 lineage artifacts are invalid")
    chdd = [
        item
        for item in artifacts
        if isinstance(item, dict) and item.get("purpose") == "canonical_chdd"
    ]
    if len(chdd) != 1 or chdd[0].get("path") != "canonical/chdd.csv":
        raise RuntimeError("selected Track 1 lineage canonical CHDD link is invalid")
    chdd_sha256 = _lower_sha256(chdd[0].get("sha256"), "selected canonical CHDD")
    chdd_path = selected / "canonical/chdd.csv"
    chdd_snapshot = _snapshot_regular_file(chdd_path, "selected canonical CHDD")
    chdd_bytes = chdd_snapshot[0]
    if sha256(chdd_bytes).hexdigest() != chdd_sha256:
        raise RuntimeError("selected canonical CHDD SHA-256 mismatch")
    snapshots = {
        lineage_path: lineage_snapshot,
        sidecar: sidecar_snapshot,
        chdd_path: chdd_snapshot,
    }
    for path, snapshot in snapshots.items():
        _assert_file_snapshot(path, snapshot, f"selected Track 1 {path.name}")
    return lineage_path, lineage_sha256, chdd_bytes, snapshots


def _result_state(value: object) -> State:
    expected_state = {"case_id", "month", "restart_ref", "wells"}
    expected_well = {
        "well",
        "role",
        "active",
        "oil_rate",
        "liquid_rate",
        "injection_rate",
        "bhp",
    }
    if not isinstance(value, dict) or set(value) != expected_state:
        raise RuntimeError("selected Track 1 state fields are invalid")
    wells = value.get("wells")
    if not isinstance(wells, list) or not wells:
        raise RuntimeError("selected Track 1 state wells are invalid")
    parsed: list[WellState] = []
    try:
        for item in wells:
            if not isinstance(item, dict) or set(item) != expected_well:
                raise RuntimeError("selected Track 1 well state fields are invalid")
            rates = (item["oil_rate"], item["liquid_rate"], item["injection_rate"])
            bhp = item["bhp"]
            if (
                not isinstance(item["well"], str)
                or not isinstance(item["role"], str)
                or not isinstance(item["active"], bool)
                or any(
                    isinstance(rate, bool) or not isinstance(rate, (int, float))
                    for rate in rates
                )
                or (
                    bhp is not None
                    and (isinstance(bhp, bool) or not isinstance(bhp, (int, float)))
                )
            ):
                raise RuntimeError("selected Track 1 well state values are invalid")
            parsed.append(
                WellState(
                    item["well"],
                    WellRole(item["role"]),
                    item["active"],
                    float(rates[0]),
                    float(rates[1]),
                    float(rates[2]),
                    None if bhp is None else float(bhp),
                )
            )
        case_id = value["case_id"]
        month = value["month"]
        restart_ref = value["restart_ref"]
        if not all(isinstance(item, str) for item in (case_id, month, restart_ref)):
            raise RuntimeError("selected Track 1 state identity is invalid")
        return State(case_id, date.fromisoformat(month), restart_ref, tuple(parsed))
    except (TypeError, ValueError) as exc:
        raise RuntimeError("selected Track 1 state contract is invalid") from exc


def _certify_selected_history(
    runner: OpmFlowRunner,
    case: Case,
    candidates: Path,
    trajectories: list[dict[str, object]],
    schedule_actions: object,
    initial_restart_ref: str,
    *,
    source: Path | None = None,
    deck: str | None = None,
    schedule_relative_path: str | None = None,
) -> tuple[State, tuple[ControlAction, ...]]:
    source = SOURCE if source is None else source
    deck = DECK if deck is None else deck
    schedule_relative_path = (
        SCHEDULE if schedule_relative_path is None else schedule_relative_path
    )
    verifier = OpmGdmBackend(
        runner,
        source,
        runs_dir=candidates,
        deck=deck,
        schedule_include=schedule_relative_path,
        normalize_model_y=True,
        parsing_strictness="low",
        source_model="Model Y",
    )
    verifier.validate_case(case)
    months = _month_range(case.start, case.end)
    if len(trajectories) != len(months):
        raise RuntimeError("selected Track 1 trajectory count differs from case horizon")
    expected_wells = set((*case.producers, *case.injectors))
    accepted: list[ControlAction] = []
    prior_restart_ref = initial_restart_ref
    final_state: State | None = None
    for month, trajectory in zip(months, trajectories, strict=True):
        if not isinstance(trajectory, dict):
            raise RuntimeError("selected Track 1 trajectory is invalid")
        state = _result_state(trajectory.get("next_state"))
        actions = verifier._actions_value(trajectory.get("actions"))
        if (
            trajectory.get("month") != month.isoformat()
            or trajectory.get("simulator") != verifier.get_provenance()
            or state.case_id != case.case_id
            or state.month != _next_month(month)
            or {well.well for well in state.wells} != expected_wells
            or len(actions) != len(expected_wells)
            or {action.well for action in actions} != expected_wells
        ):
            raise RuntimeError("selected Track 1 state/action continuity is invalid")
        accepted.extend(actions)
        history = verifier._authenticated_history(case, state)
        if history != tuple(accepted):
            raise RuntimeError("selected Track 1 lineage history differs from result")
        lineage_path, lineage_sha256 = verifier._parse_restart_ref(state.restart_ref)
        lineage, _ = _authenticated_json(
            lineage_path, lineage_sha256, "selected Track 1 step lineage"
        )
        if (
            lineage.get("run_id") != trajectory.get("run_id")
            or lineage.get("prior_restart_ref") != prior_restart_ref
        ):
            raise RuntimeError("selected Track 1 restart lineage is discontinuous")
        prior_restart_ref = state.restart_ref
        final_state = state

    compiled = ScheduleCompiler().validate(case, accepted)
    if compiled != verifier._actions_value(schedule_actions):
        raise RuntimeError("selected Track 1 schedule differs from certified history")
    assert final_state is not None
    return final_state, compiled


def _records(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as stream:
        return list(csv.DictReader(stream))


def _organizer_reference(path: Path | None = None) -> dict[str, object]:
    path = ORGANIZER_REFERENCE if path is None else path
    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        sheet = workbook.active
        years = [int(sheet.cell(1, column).value) for column in range(3, 12)]
        rows = {
            str(sheet.cell(row, 1).value): [
                float(sheet.cell(row, column).value or 0)
                for column in range(3, 12)
            ]
            for row in range(2, sheet.max_row + 1)
        }
    finally:
        workbook.close()
    required = {
        "Добыча нефти",
        "Добыча жидкости",
        "Операции по смене ЭЦН",
        "Количество установленных новых насосов",
        "FCF",
        "Коэффициент дисконтирования",
    }
    if not required <= rows.keys():
        raise RuntimeError(
            f"organizer workbook misses rows: {sorted(required - rows.keys())}"
        )
    return {
        "annual_mass_kt": {
            str(year): {
                "oil": rows["Добыча нефти"][index],
                "liquid": rows["Добыча жидкости"][index],
            }
            for index, year in enumerate(years)
        },
        "pump_changes": sum(rows["Количество установленных новых насосов"]),
        "pump_operation_m": sum(rows["Операции по смене ЭЦН"]),
        "rebased_chdd_2014_m": sum(
            rows["FCF"][index]
            * rows["Коэффициент дисконтирования"][index]
            / rows["Коэффициент дисконтирования"][years.index(2014)]
            for index, year in enumerate(years)
            if year >= 2014
        ),
    }


def _assert_organizer_reference_parity(
    chdd_records: list[dict[str, str]],
    reference_result_dir: Path,
    reference_workbook: Path | None = None,
) -> dict[str, object]:
    reference_workbook = (
        ORGANIZER_REFERENCE if reference_workbook is None else reference_workbook
    )
    reference = _organizer_reference(reference_workbook)
    expected_mass = reference["annual_mass_kt"]
    assert isinstance(expected_mass, dict)
    actual_mass: dict[str, dict[str, float]] = {}
    for row in chdd_records:
        year = row["DATA"][:4]
        annual = actual_mass.setdefault(year, {"oil": 0.0, "liquid": 0.0})
        annual["oil"] += float(row["WOMT_Diff"]) / 1000
        annual["liquid"] += float(row["WLPT_Diff"]) / 1000
    if actual_mass.keys() != expected_mass.keys():
        raise RuntimeError(
            "canonical and organizer annual years differ: "
            f"actual={sorted(actual_mass)}, expected={sorted(expected_mass)}"
        )
    annual_evidence: dict[str, object] = {}
    for year, expected in expected_mass.items():
        assert isinstance(expected, dict)
        actual = actual_mass[year]
        residual = {
            phase: actual[phase] - float(expected[phase])
            for phase in ("oil", "liquid")
        }
        tolerances = {
            phase: 2**-22 * max(1.0, abs(float(expected[phase])))
            for phase in ("oil", "liquid")
        }
        if any(abs(residual[phase]) > tolerances[phase] for phase in residual):
            raise RuntimeError(
                f"organizer annual mass parity failed for {year}: residual={residual} "
                f"kt, float32_vector_tolerance={tolerances} kt"
            )
        annual_evidence[year] = {
            "expected_kt": expected,
            "actual_kt": actual,
            "residual_kt": residual,
            "float32_vector_tolerance_kt": tolerances,
        }

    raw_result = json.loads((reference_result_dir / "result.json").read_text())
    pump_changes = float(raw_result["summary"]["pumpChanges"])
    pump_operation_m = sum(
        float(annual["pumpOperationM"]) for annual in raw_result["annual"]
    )
    if pump_changes != float(reference["pump_changes"]):
        raise RuntimeError(
            f"organizer pumpChanges parity failed: {pump_changes} != "
            f"{reference['pump_changes']}"
        )
    if not isclose(
        pump_operation_m,
        float(reference["pump_operation_m"]),
        rel_tol=0,
        abs_tol=1e-9,
    ):
        raise RuntimeError(
            f"organizer pumpOperationM parity failed: {pump_operation_m} != "
            f"{reference['pump_operation_m']}"
        )
    rebased_2014 = float(reference["rebased_chdd_2014_m"])
    if not isclose(rebased_2014, TARGETS[2014], rel_tol=0, abs_tol=1e-9):
        raise RuntimeError(
            f"organizer rebased 2014 CHDD differs from registered target: "
            f"{rebased_2014} != {TARGETS[2014]}"
        )
    return {
        "workbook": str(reference_workbook),
        "workbook_sha256": _sha256_file(reference_workbook),
        "annual_mass": annual_evidence,
        "annual_mass_tolerance": "2^-22 * max(1 kt, abs(reference annual mass))",
        "pump_changes": pump_changes,
        "pump_operation_m": pump_operation_m,
        "rebased_chdd_2014_m": rebased_2014,
    }


def _action(row: dict[str, str]) -> ControlAction:
    target = ControlTarget(row["control_target"])
    role = (
        WellRole.INJECTOR
        if target is ControlTarget.WATER_INJECTION_RATE
        else WellRole.PRODUCER
    )
    status = WellStatus.OPEN if int(row["status"]) == 1 else WellStatus.SHUT
    return ControlAction(
        date.fromisoformat(row["date"]),
        row["well"],
        role,
        status,
        target,
        float(row["control_value"]) if status is WellStatus.OPEN else 0.0,
    )


def _action_json(action: ControlAction) -> dict[str, object]:
    return {
        "well": action.well,
        "role": action.role.value,
        "status": action.status.value,
        "target": action.target.value,
        "value": action.value,
    }


def _state_json(state: State) -> dict[str, object]:
    return {
        "case_id": state.case_id,
        "month": state.month.isoformat(),
        "restart_ref": state.restart_ref,
        "wells": [
            {
                "well": well.well,
                "role": well.role.value,
                "active": well.active,
                "oil_rate": well.oil_rate,
                "liquid_rate": well.liquid_rate,
                "injection_rate": well.injection_rate,
                "bhp": well.bhp,
            }
            for well in state.wells
        ],
    }


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--resume-baseline", action="store_true")
    parser.add_argument("--output-root", type=Path)
    parser.add_argument("--source", type=Path, default=SOURCE)
    parser.add_argument("--baseline-dir", type=Path, default=BASELINE)
    parser.add_argument(
        "--reference-workbook", type=Path, default=ORGANIZER_REFERENCE
    )
    parser.add_argument("--deck", default=DECK)
    parser.add_argument("--schedule-relative-path", default=SCHEDULE)
    args = parser.parse_args(argv)
    source = _runtime_regular_file(args.source, "Model Y source")
    reference_workbook = _runtime_regular_file(
        args.reference_workbook, "organizer reference workbook"
    )
    baseline_dir = args.baseline_dir.absolute()
    _reject_symlink_components(baseline_dir)
    if args.resume_baseline and (
        baseline_dir.resolve() != baseline_dir
        or baseline_dir.is_symlink()
        or not baseline_dir.is_dir()
    ):
        raise RuntimeError(
            f"baseline directory is not a regular non-symlink directory: {baseline_dir}"
        )
    deck = _safe_relative_path(args.deck, "Model Y deck")
    schedule_relative_path = _safe_relative_path(
        args.schedule_relative_path, "Model Y schedule"
    )
    baseline, candidates, final, config_path, artifact_root = _output_paths(
        args.output_root,
        resume_baseline=args.resume_baseline,
        source=source,
        baseline_dir=baseline_dir,
        reference_workbook=reference_workbook,
    )
    proof_root = Path(__file__).absolute().parents[1]
    controller = Path(__file__).absolute().with_name("run_track1_mpc.py")
    source_contract = _proof_source_contract(controller)
    _assert_no_running_opm()
    if _sha256_file(source) != SOURCE_SHA256:
        raise RuntimeError("Model Y source SHA-256 mismatch")
    runner = OpmFlowRunner(timeout_seconds=3600)
    if args.resume_baseline:
        result, report, extraction = _resume_baseline(
            runner,
            baseline=baseline,
            candidates=candidates,
            source=source,
            deck=deck,
            schedule_relative_path=schedule_relative_path,
        )
    else:
        result = runner.run(
            source,
            baseline,
            deck=deck,
            parsing_strictness="low",
            normalize_model_y=True,
        )
        report, extraction = runner.extract_summary_report(
            result, result.run_dir / "summary-report.txt"
        )
    run_manifest, _ = _authenticated_json(
        result.manifest_path, result.manifest_sha256, "OPM run manifest"
    )
    vectors = _validated_manifest_vectors(run_manifest.get("summary_contract"))
    if not MASS_VECTORS <= vectors:
        raise RuntimeError(
            f"expanded multi-PVT summary contract is missing {sorted(MASS_VECTORS - vectors)}"
        )
    canonical = artifact_root / "canonical"
    chdd_csv = canonical / "chdd.csv"
    trajectory_csv = canonical / "trajectory.csv"
    export_manifest = canonical / "manifest.json"
    export_opm_chdd(
        report,
        chdd_csv,
        trajectory_csv,
        export_manifest,
        scenario_id="model-y-baseline-20260831-a100-v4",
        source_model="Model Y",
        opm_run_manifest=result.manifest_path,
        summary_extraction_manifest=extraction,
        deck_dir=result.deck_path.parent,
        unit_system="METRIC",
    )
    _authenticated_json(
        result.manifest_path, result.manifest_sha256, "OPM run manifest"
    )
    adapter = CHDDEconomicsAdapter()
    chdd_records = _records(chdd_csv)
    economics = {
        year: {
            "canonical": adapter.calculate(
                chdd_records,
                start_year=year,
                output_dir=artifact_root / f"economics-{year}-canonical",
                charge_initial_pump=False,
            ),
            "organizer_reference": adapter.calculate(
                chdd_records,
                start_year=year,
                output_dir=artifact_root / f"economics-{year}-organizer-reference",
                charge_initial_pump=True,
            ),
        }
        for year in TARGETS
    }
    chdd_parity: dict[int, dict[str, float]] = {}
    for year, target in TARGETS.items():
        target_profile = TARGET_PROFILES[year]
        actual = economics[year][target_profile].total_chdd_m
        residual = actual - target
        tolerance = (
            CHDD_ABS_TOLERANCE_M
            if year == 2007
            else 2**-22 * max(1.0, abs(target))
        )
        if abs(residual) > tolerance:
            raise RuntimeError(
                f"Model Y CHDD parity failed for {year}: actual={actual:.9f}, "
                f"target={target:.9f}, abs_tolerance={tolerance:.9f}"
            )
        chdd_parity[year] = {
            "target_chdd_m": target,
            "actual_chdd_m": actual,
            "residual_chdd_m": residual,
            "abs_tolerance_chdd_m": tolerance,
        }
    organizer_parity = _assert_organizer_reference_parity(
        chdd_records,
        economics[2007]["organizer_reference"].output_dir,
        reference_workbook,
    )
    horizon_report_end = _next_month(HORIZON_END)
    baseline_horizon_records = [
        record
        for record in chdd_records
        if HORIZON_START
        <= date.fromisoformat(record["DATA"])
        <= horizon_report_end
    ]
    horizon_dates = sorted(
        {date.fromisoformat(record["DATA"]) for record in baseline_horizon_records}
    )
    if not horizon_dates or horizon_dates[0] != HORIZON_START or horizon_dates[-1] != horizon_report_end:
        raise RuntimeError("baseline CHDD rows do not cover the six-month Track 1 window")
    baseline_horizon_economics = adapter.calculate(
        baseline_horizon_records,
        start_year=HORIZON_START.year,
        output_dir=artifact_root / "economics-2014-six-month-baseline",
        charge_initial_pump=False,
    )
    baseline_proof = {
        "schema": "timesoil.aios.model-y-baseline-proof/v1",
        "script_source_contract": source_contract,
        "source_sha256": SOURCE_SHA256,
        "opm_run_manifest": str(result.manifest_path),
        "opm_run_manifest_sha256": result.manifest_sha256,
        "summary_report_sha256": _sha256_file(report),
        "summary_extraction_sha256": _sha256_file(extraction),
        "canonical_export_manifest_sha256": _sha256_file(export_manifest),
        "canonical_chdd_sha256": _sha256_file(chdd_csv),
        "six_month_horizon": {
            "start": HORIZON_START.isoformat(),
            "end": HORIZON_END.isoformat(),
            "report_through": horizon_report_end.isoformat(),
            "chdd_m": baseline_horizon_economics.total_chdd_m,
            "economics_manifest": str(baseline_horizon_economics.manifest_path),
            "economics_manifest_sha256": _sha256_file(
                baseline_horizon_economics.manifest_path
            ),
        },
        "organizer_reference_parity": organizer_parity,
        "simulator_envelope_2007_chdd_m": {
            "tnavigator": SIMULATOR_ENVELOPE_2007[0],
            "opm_flow": SIMULATOR_ENVELOPE_2007[1],
            "absolute_gap": SIMULATOR_ENVELOPE_2007[1]
            - SIMULATOR_ENVELOPE_2007[0],
        },
        "economics": {
            str(year): {
                **chdd_parity[year],
                "target_profile": TARGET_PROFILES[year],
                "target_semantics": (
                    "initial field pump deployment charged"
                    if year == 2007
                    else "rebased full-run FCF; pre-2014 pumps are sunk assets"
                ),
                "profiles": {
                    profile: {
                        "charge_initial_pump": profile == "organizer_reference",
                        "actual_chdd_m": calculation.total_chdd_m,
                        "start_date": calculation.start_date,
                        "max_date": calculation.max_date,
                        "diagnostics": dict(calculation.diagnostics),
                        "manifest": str(calculation.manifest_path),
                        "manifest_sha256": _sha256_file(calculation.manifest_path),
                    }
                    for profile, calculation in economics[year].items()
                },
            }
            for year in TARGETS
        },
    }
    baseline_proof_path = artifact_root / "baseline-proof.json"
    _write_new_json(baseline_proof_path, baseline_proof)

    months = _month_range(HORIZON_START, HORIZON_END)
    if len(months) != HORIZON_MONTH_COUNT:
        raise RuntimeError("Track 1 proof horizon is not exactly six months")
    trajectory_rows = _records(trajectory_csv)
    baseline_by_month = {
        month: tuple(
            sorted(
                (
                    _action(row)
                    for row in trajectory_rows
                    if row["date"] == month.isoformat()
                ),
                key=lambda action: (
                    action.role.value,
                    action.well,
                    action.status.value,
                    action.target.value,
                    action.value,
                ),
            )
        )
        for month in months
    }
    if any(len(actions) != 49 for actions in baseline_by_month.values()):
        counts = {month.isoformat(): len(actions) for month, actions in baseline_by_month.items()}
        raise RuntimeError(f"expected 49 controls in every Track 1 month, got {counts}")
    first_actions = baseline_by_month[months[0]]
    producers = tuple(
        action.well for action in first_actions if action.role is WellRole.PRODUCER
    )
    injectors = tuple(
        action.well for action in first_actions if action.role is WellRole.INJECTOR
    )
    if len(producers) != 33 or len(injectors) != 16:
        raise RuntimeError("Model Y well roles differ from 33 producers/16 injectors")
    case = Case(
        "model-y-track1-proof",
        months[0],
        months[-1],
        months[0],
        producers,
        injectors,
    )
    compiler = ScheduleCompiler()
    expected_wells = set((*producers, *injectors))
    for month, actions in baseline_by_month.items():
        if {action.well for action in actions} != expected_wells:
            raise RuntimeError(f"Model Y well set changed in {month.isoformat()}")
        compiler.validate(case, actions)
    baseline_actions = tuple(
        action for month in months for action in baseline_by_month[month]
    )
    baseline_schedule = compiler.compile(case, baseline_actions)
    baseline_ref = (
        f"{result.manifest_path.resolve()}#sha256={result.manifest_sha256}"
    )
    state = State(
        case.case_id,
        case.start,
        baseline_ref,
        OpmGdmBackend._wells_at(
            report, case, case.start, deck_dir=result.deck_path.parent
        ),
    )
    candidates_by_month: dict[date, tuple[tuple[ControlAction, ...], ...]] = {}
    for month, monthly_baseline in baseline_by_month.items():
        adjustable = next(
            (
                action
                for action in monthly_baseline
                if action.well == "2"
                and action.role is WellRole.INJECTOR
                and action.status is WellStatus.OPEN
                and action.value >= CONTROL_PERTURBATION
            ),
            None,
        )
        if adjustable is None:
            raise RuntimeError(
                f"bounded injector-2 candidates cannot be built for {month.isoformat()}"
            )
        lower = tuple(
            replace(action, value=action.value - CONTROL_PERTURBATION)
            if action is adjustable
            else action
            for action in monthly_baseline
        )
        upper = tuple(
            replace(action, value=action.value + CONTROL_PERTURBATION)
            if action is adjustable
            else action
            for action in monthly_baseline
        )
        variants = (monthly_baseline, lower, upper)
        if len(set(variants)) != 3:
            raise RuntimeError("safe bounded candidates are not distinct")
        for variant in variants:
            compiler.validate(case, variant)
        candidates_by_month[month] = variants
    config = {
        "schema": "timesoil.aios.track1-mpc-input/v1",
        "case": {
            "case_id": case.case_id,
            "start": case.start.isoformat(),
            "end": case.end.isoformat(),
            "economics_start": case.economics_start.isoformat(),
            "producers": list(case.producers),
            "injectors": list(case.injectors),
            "max_liquid_rate": case.max_liquid_rate,
        },
        "initial_state": _state_json(state),
        "candidates": {
            month.isoformat(): [
                [_action_json(action) for action in candidate]
                for candidate in candidates_by_month[month]
            ]
            for month in months
        },
        "opm": {
            "source": str(source),
            "runs_dir": str(candidates),
            "deck": deck,
            "schedule_include": schedule_relative_path,
            "normalize_model_y": True,
            "parsing_strictness": "low",
            "source_model": "Model Y",
            "timeout_seconds": 3600,
        },
    }
    _write_new_json(config_path, config)
    subprocess.run(
        [
            sys.executable,
            str(controller),
            str(config_path),
            "--runs-dir",
            str(final),
            "--proof-script",
            str(Path(__file__).absolute()),
        ],
        cwd=proof_root,
        check=True,
        stdin=subprocess.DEVNULL,
        shell=False,
    )
    _verify_proof_source_contract(source_contract)

    final_runs = [path for path in final.iterdir() if path.is_dir()]
    if len(final_runs) != 1:
        raise RuntimeError("Track 1 CLI did not produce exactly one result directory")
    final_run = final_runs[0]
    (
        result_value,
        result_manifest_sha256,
        schedule_sha256,
        run_snapshots,
    ) = _verify_track1_run(final_run, source_contract["execution"])
    config_sha256 = _sha256_file(config_path)
    expected_input_sha256 = sha256(
        (
            json.dumps(
                {
                    "config_sha256": config_sha256,
                    "source_sha256": SOURCE_SHA256,
                },
                ensure_ascii=False,
                sort_keys=True,
                indent=2,
                allow_nan=False,
            )
            + "\n"
        ).encode()
    ).hexdigest()
    case_value = config["case"]
    if (
        not isinstance(case_value, dict)
        or result_value.get("case_id") != case_value.get("case_id")
        or result_value.get("config_sha256") != config_sha256
        or result_value.get("source_sha256") != SOURCE_SHA256
        or result_value.get("input_sha256") != expected_input_sha256
        or _sha256_file(source) != SOURCE_SHA256
    ):
        raise RuntimeError("Track 1 result differs from proof config or Model Y source")
    evidence = result_value["evidence"]
    if not isinstance(evidence, dict):
        raise RuntimeError("Track 1 result evidence is invalid")
    trajectories = evidence["trajectories"]
    if (
        not isinstance(trajectories, list)
        or len(trajectories) != HORIZON_MONTH_COUNT
        or not all(isinstance(trajectory, dict) for trajectory in trajectories)
    ):
        raise RuntimeError("Track 1 proof requires exactly six certified trajectories")
    terminal_trajectory = trajectories[-1]
    assert isinstance(terminal_trajectory, dict)
    selected_run_id = terminal_trajectory["run_id"]
    next_state = terminal_trajectory.get("next_state")
    if not isinstance(next_state, dict):
        raise RuntimeError("selected Track 1 next_state is invalid")
    (
        lineage_path,
        lineage_sha256,
        selected_chdd_bytes,
        selected_snapshots,
    ) = _verified_selected_lineage(
        candidates, selected_run_id, next_state.get("restart_ref")
    )
    schedule_value = result_value.get("schedule")
    if not isinstance(schedule_value, dict):
        raise RuntimeError("selected Track 1 schedule is invalid")
    certified_state, certified_history = _certify_selected_history(
        runner,
        case,
        candidates,
        trajectories,
        schedule_value.get("actions"),
        baseline_ref,
        source=source,
        deck=deck,
        schedule_relative_path=schedule_relative_path,
    )
    for path, snapshot in selected_snapshots.items():
        _assert_file_snapshot(path, snapshot, f"selected Track 1 {path.name}")
    copied_chdd = final_run / "selected_chdd.csv"
    with copied_chdd.open("xb") as target:
        target.write(selected_chdd_bytes)
    copied_chdd.chmod(0o444)
    _assert_file_bytes(copied_chdd, selected_chdd_bytes, "copied selected CHDD")
    baseline_schedule_path = final_run / "baseline_wells_schedule.inc"
    baseline_schedule_bytes = baseline_schedule.text.encode("utf-8")
    with baseline_schedule_path.open("xb") as target:
        target.write(baseline_schedule_bytes)
    baseline_schedule_path.chmod(0o444)
    _assert_file_bytes(
        baseline_schedule_path, baseline_schedule_bytes, "copied baseline schedule"
    )
    economics = evidence["step_economics"]
    if (
        not isinstance(economics, list)
        or len(economics) != HORIZON_MONTH_COUNT
        or not all(isinstance(item, dict) for item in economics)
    ):
        raise RuntimeError("Track 1 proof requires exactly six economics results")
    terminal_economics = economics[-1]
    assert isinstance(terminal_economics, dict)
    selected_chdd_m = terminal_economics.get("npv_million_rub")
    lineage_value, _ = _authenticated_json(
        lineage_path, lineage_sha256, "selected Track 1 terminal lineage"
    )
    lineage_economics = lineage_value.get("economics")
    lineage_chdd_m = (
        lineage_economics.get("total_chdd_m")
        if isinstance(lineage_economics, dict)
        else None
    )
    if (
        isinstance(selected_chdd_m, bool)
        or not isinstance(selected_chdd_m, (int, float))
        or isinstance(lineage_chdd_m, bool)
        or not isinstance(lineage_chdd_m, (int, float))
        or not isclose(
            float(selected_chdd_m),
            float(lineage_chdd_m),
            rel_tol=0.0,
            abs_tol=1e-9,
        )
    ):
        raise RuntimeError("selected terminal CHDD differs from authenticated lineage")
    baseline_chdd_m = baseline_horizon_economics.total_chdd_m
    proof = {
        "schema": "timesoil.aios.model-y-track1-proof/v2",
        "certified": True,
        "script_source_contract": source_contract,
        "replay_mode": "full-replay",
        "binary_restart": False,
        "horizon": {
            "start": case.start.isoformat(),
            "end": case.end.isoformat(),
            "report_through": certified_state.month.isoformat(),
            "month_count": len(months),
            "well_count": len(expected_wells),
            "producer_count": len(producers),
            "injector_count": len(injectors),
            "candidate_count_per_month": 3,
        },
        "baseline_proof": str(baseline_proof_path),
        "baseline_proof_sha256": _sha256_file(baseline_proof_path),
        "baseline": {
            "opm_run_manifest": str(result.manifest_path),
            "opm_run_manifest_sha256": result.manifest_sha256,
            "schedule": str(baseline_schedule_path),
            "schedule_sha256": baseline_schedule.sha256,
            "chdd_m": baseline_chdd_m,
            "economics_manifest": str(baseline_horizon_economics.manifest_path),
            "economics_manifest_sha256": _sha256_file(
                baseline_horizon_economics.manifest_path
            ),
        },
        "config": str(config_path),
        "config_sha256": config_sha256,
        "result_manifest_sha256": result_manifest_sha256,
        "wells_schedule_sha256": schedule_sha256,
        "selected_run_id": selected_run_id,
        "selected_chdd_m": selected_chdd_m,
        "selected_chdd_sha256": sha256(selected_chdd_bytes).hexdigest(),
        "selected_lineage": str(lineage_path),
        "selected_lineage_sha256": lineage_sha256,
        "selected": {
            "terminal_run_id": selected_run_id,
            "terminal_lineage": str(lineage_path),
            "terminal_lineage_sha256": lineage_sha256,
            "schedule": str(final_run / "wells_schedule.inc"),
            "schedule_sha256": schedule_sha256,
            "canonical_chdd": str(copied_chdd),
            "canonical_chdd_sha256": sha256(selected_chdd_bytes).hexdigest(),
            "chdd_m": selected_chdd_m,
        },
        "comparison": {
            "selected_minus_baseline_chdd_m": float(selected_chdd_m)
            - baseline_chdd_m,
            "selected_schedule_differs": certified_history
            != baseline_schedule.actions,
        },
        "candidate_run_ids": [
            path.name for path in sorted(candidates.iterdir()) if path.is_dir()
        ],
    }
    _verify_proof_source_contract(source_contract)
    if _sha256_file(config_path) != config_sha256 or _sha256_file(source) != SOURCE_SHA256:
        raise RuntimeError("Track 1 proof config or Model Y source changed before receipt")
    (
        verified_result,
        verified_manifest_sha256,
        verified_schedule_sha256,
        verified_run_snapshots,
    ) = _verify_track1_run(final_run, source_contract["execution"])
    (
        verified_lineage,
        verified_lineage_sha256,
        verified_chdd,
        verified_selected_snapshots,
    ) = _verified_selected_lineage(
        candidates, selected_run_id, next_state["restart_ref"]
    )
    verified_state, verified_history = _certify_selected_history(
        runner,
        case,
        candidates,
        trajectories,
        schedule_value.get("actions"),
        baseline_ref,
        source=source,
        deck=deck,
        schedule_relative_path=schedule_relative_path,
    )
    if (
        verified_result != result_value
        or verified_manifest_sha256 != result_manifest_sha256
        or verified_schedule_sha256 != schedule_sha256
        or verified_lineage != lineage_path
        or verified_lineage_sha256 != lineage_sha256
        or verified_chdd != selected_chdd_bytes
        or verified_run_snapshots != run_snapshots
        or verified_selected_snapshots != selected_snapshots
        or verified_state != certified_state
        or verified_history != certified_history
    ):
        raise RuntimeError("Track 1 authenticated artifacts changed before receipt")
    for path, snapshot in (*run_snapshots.items(), *selected_snapshots.items()):
        _assert_file_snapshot(path, snapshot, f"Track 1 receipt source {path.name}")
    _assert_file_bytes(copied_chdd, selected_chdd_bytes, "copied selected CHDD")
    _assert_file_bytes(
        baseline_schedule_path, baseline_schedule_bytes, "copied baseline schedule"
    )
    _write_new_json(final_run / "proof.json", proof)
    (final_run / "proof.json").chmod(0o444)
    print(json.dumps(proof, ensure_ascii=False, sort_keys=True))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
